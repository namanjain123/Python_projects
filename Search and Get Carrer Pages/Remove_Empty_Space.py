# import openpyxl library

import openpyxl
# function to remove empty rows
def remove(sheet, row):
    # iterate the row object
    for cell in row:
        # check the value of each cell in
        # the row, if any of the value is not
        # None return without removing the row
        if cell.value != None:
              return
    # get the row number from the first cell
    # and remove the row
    sheet.delete_rows(row[0].row, 1)
if __name__ == '__main__':
    # enter your file path
    path = 'A:\Python Projects\Search and Get Carrer Pages\Company List.xlsx'
    # load excel file
    book = openpyxl.load_workbook(path)
    # select the sheet
    sheet = book['Sheet1']
    print("Maximum rows before removing:", sheet.max_row)
    # iterate the sheet object
    for row in sheet:
      remove(sheet,row)
    print("Maximum rows after removing:",sheet.max_row)
    # save the file to the path
    path='A:\Python Projects\Search and Get Carrer Pages\Company List.xlsx'
    book.save(path)